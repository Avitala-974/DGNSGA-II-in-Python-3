import DCNSGA_II_DE_tools
import DCNSGA_II_DE_conf
import dynamic_tools
import copy
import os
import nichec
import sys
from functools import cmp_to_key
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
import numpy as np

# 设置工作目录、问题库和结果输出目录
WORKING_DIR = os.getcwd()
PROBLEM_DIR = os.path.join(WORKING_DIR, "PROBLEM")
RESULT_DIR = os.path.join(WORKING_DIR, "RESULT")
PARETO_DIR = os.path.join(RESULT_DIR, "Pareto Solutions")  # 帕累托解存储目录
LOCAL_PATH = [WORKING_DIR, PROBLEM_DIR, RESULT_DIR]
sys.path.extend(LOCAL_PATH)

# 创建结果目录
if not os.path.exists(RESULT_DIR):
    os.makedirs(RESULT_DIR)
if not os.path.exists(PARETO_DIR):
    os.makedirs(PARETO_DIR)


def visualize_gene_structure(individual, num_provinces=3, num_years=2):
    """可视化个体的基因结构，展示前N个省份和前M年的基因编码规则"""
    genes = individual['genes']
    print(f"\n===== 基因结构可视化（总长度: {len(genes)}） =====")

    # 基因编码规则说明
    print("基因编码规则（每省份每年6个基因）:")
    print("  [0: 第一产业增长率, 1: 第二产业增长率, 2: 第三产业增长率, "
          "3: 煤炭消费量, 4: 石油消费量, 5: 天然气消费量]")

    for prov in range(num_provinces):
        print(f"\n----- 省份 {prov} 的基因片段 -----")
        for year in range(num_years):
            # 计算该省份该年份的基因起始索引
            base_idx = prov * 48 + year * 6  # 每省份48个基因（8年×6）
            if base_idx + 6 > len(genes):
                break  # 避免索引越界

            # 提取6个基因值
            year_genes = genes[base_idx: base_idx + 6]
            growth_rates = year_genes[:3]  # 前3个：增长率
            energy = year_genes[3:6]  # 后3个：能源消费（煤、油、气）

            # 打印详细信息
            print(f"年份 {2023 + year}（索引 {base_idx}-{base_idx + 5}）:")
            print(f"  产业增长率: 一产={growth_rates[0]:.4f}, 二产={growth_rates[1]:.4f}, 三产={growth_rates[2]:.4f}")
            print(f"  能源消费量: 煤炭={energy[0]:.2f}, 石油={energy[1]:.2f}, 天然气={energy[2]:.2f}")
            print(f"  能源总量: {sum(energy):.2f}")

    print("\n==========================================")


def verify_energy_calculation(individual, province=0, year=2025):
    """验证特定省份和年份的能源计算是否正确"""
    genes = individual['genes']
    year_idx = year - 2023  # 2023→0, 2024→1, ..., 2030→7
    if year_idx < 0 or year_idx > 7:
        print(f"错误：年份 {year} 超出范围（2023-2030）")
        return

    # 计算该省份该年份的基因索引
    base_idx = province * 48 + year_idx * 6
    if base_idx + 6 > len(genes):
        print(f"错误：基因索引超出范围（省份 {province}, 年份 {year}）")
        return

    # 提取能源数据
    coal = genes[base_idx + 3]
    oil = genes[base_idx + 4]
    gas = genes[base_idx + 5]
    total = coal + oil + gas

    print(f"\n===== 能源计算验证（省份 {province}, 年份 {year}） =====")
    print(f"基因索引: {base_idx + 3}（煤）, {base_idx + 4}（油）, {base_idx + 5}（气）")
    print(f"数值: 煤炭={coal:.2f}, 石油={oil:.2f}, 天然气={gas:.2f}")
    print(f"总和: {total:.2f}")
    return total


def init(popSize, problem_initialize, evaluator):
    """初始化算法"""
    global parent_size, offspring_size, _genCount, _evaluator, parent_pop, upper, lower, constraints_num, objectives_number, evaluationTime
    parent_size, offspring_size, _genCount, _evaluator, upper, lower = (
        popSize, popSize,
        problem_initialize[0], evaluator,
        problem_initialize[1], problem_initialize[2]
    )
    constraints_num, objectives_number = problem_initialize[4], problem_initialize[5]
    parent_pop = dynamic_tools.initialize_parent_population(parent_size, _genCount)
    dynamic_tools.caculate_pheno(parent_pop, upper, lower, _genCount, parent_size)

    # 初始化种群后，可视化第一个个体的基因结构
    if parent_pop:  # 确保种群非空
        print("\n===== 初始种群基因结构分析 =====")
        visualize_gene_structure(parent_pop[0])  # 可视化第一个个体
        # 验证2025年能源计算
        verify_energy_calculation(parent_pop[0], province=0, year=2025)
        verify_energy_calculation(parent_pop[0], province=0, year=2030)

    evaluationTime = dynamic_tools.evaluate_population(parent_pop, _evaluator, dynamic_tools.get_fill_result)


def loop(generation, outputfreq, condition):
    """主循环"""
    global parent_pop, evaluationTime
    initialMaxViolation = dynamic_tools.caculate_initial_max_violation(parent_pop)
    e = initialMaxViolation
    dynamic_tools.caculate_violation_objective(initialMaxViolation, parent_pop)
    dynamic_tools.mark_individual_efeasible(e, parent_pop)
    K, g = 0, 0
    MaxK = DCNSGA_II_DE_conf.MaxK
    normalized_upper, normalized_lower = [1.0] * _genCount, [0.0] * _genCount
    R = nichec.get_MaxR(_genCount, parent_size + offspring_size, normalized_upper, normalized_lower)

    print(" 初始最大违反约束:", ", ".join(f"{v:.4e}" for v in initialMaxViolation))
    print(f" 初始化R: {R:.4e}, MaxK: {MaxK}")

    while K <= MaxK:
        feasible_ratio = dynamic_tools.get_efeasible_ratio(parent_pop)
        print(f"\n📘 Generation {g}, State K = {K}, 可行解比例: {feasible_ratio:.2%}")
        print("🔧 当前 e 向量 = [" + ", ".join(f"{ei:.4e}" for ei in e) + "]")
        # 每10代可视化一次最优个体的基因结构（避免输出过多）
        if g % 10 == 0 and parent_pop:
            print("\n===== 当代最优个体基因结构分析 =====")
            # 取违反值最小的个体
            best_ind = min(parent_pop, key=lambda x: sum(x['violations']))
            visualize_gene_structure(best_ind, num_provinces=1, num_years=3)  # 简化输出
            # 验证修复后的能源计算
            verify_energy_calculation(best_ind, province=0, year=2025)
            verify_energy_calculation(best_ind, province=0, year=2030)

        # 当可行解比例大于80%时，更新 K 和 e
        if feasible_ratio > 0.9:
            print("✅ 可行解比例超过90%，进入状态提升（K += 1）")
            K += 1
            if K > MaxK:
                break
            e = dynamic_tools.reduce_boundary(initialMaxViolation, K, MaxK)
            r = nichec.reduce_radius(K, MaxK, _genCount, R, upper, lower)
            print(f"🔧 e[0]更新为: {e[0]:.4e}, r 更新为: {r:.4e}")
            dynamic_tools.mark_individual_efeasible(e, parent_pop)
        else:
            print("⛔ 可行解比例不足90%，K 保持不变")

        offspring_pop = dynamic_tools.generate_offspring_population(g, offspring_size, parent_pop, _genCount)
        dynamic_tools.caculate_pheno(offspring_pop, upper, lower, _genCount, offspring_size)
        evaluationTime += dynamic_tools.evaluate_population(offspring_pop, _evaluator, dynamic_tools.get_fill_result)
        dynamic_tools.caculate_violation_objective(initialMaxViolation, offspring_pop)
        dynamic_tools.mark_individual_efeasible(e, offspring_pop)

        nichec.caculate_nichecount(parent_pop, offspring_pop, _genCount, r, parent_size + offspring_size)
        parent_pop = DCNSGA_II_DE_tools.select_next_parent_population(offspring_pop, parent_pop, parent_size)

        if g == generation:
            break
        g += 1

    parent_pop.sort(key=cmp_to_key(compare))
    bestObj = parent_pop[0]
    return bestObj, evaluationTime, g, parent_pop  # 返回完整的父代种群用于Excel输出


def compare(a, b):
    """个体排序比较函数"""
    if a['violation_objectives'][0] < b['violation_objectives'][0]: return -1
    if a['violation_objectives'][0] > b['violation_objectives'][0]: return 1
    if a['objectives'] < b['objectives']: return -1
    if a['objectives'] > b['objectives']: return 1
    return 0


def run(problem_initialize, generation, popsize, evaluator, outputfreq=1, condition=lambda x: False):
    init(popsize, problem_initialize, evaluator)
    return loop(generation, outputfreq, condition)


def get_average(res):
    return sum(res) / float(len(res))


def get_variance(res, ave):
    return pow(sum((float(x) - ave) ** 2 for x in res) / len(res), 0.5)


def is_pareto_efficient(costs):
    """
    找出帕累托有效点
    参考: https://stackoverflow.com/questions/32791911/fast-calculation-of-pareto-front-in-python
    """
    is_efficient = np.ones(costs.shape[0], dtype=bool)
    for i, c in enumerate(costs):
        if is_efficient[i]:
            is_efficient[is_efficient] = np.any(costs[is_efficient] < c, axis=1)
            is_efficient[i] = True
    return is_efficient


def save_pareto_solutions(population, problem_module):
    """保存所有帕累托前沿解到独立的Excel文件"""
    # 提取所有个体的目标函数值
    objectives = np.array([ind['objectives'] for ind in population])

    # 找出帕累托有效解
    pareto_mask = is_pareto_efficient(objectives)
    pareto_solutions = [population[i] for i in range(len(population)) if pareto_mask[i]]

    print(f"\n发现 {len(pareto_solutions)} 个帕累托前沿解，正在保存...")

    # 为每个帕累托解生成Excel文件
    pareto_data = []
    for i, sol in enumerate(pareto_solutions):
        # 计算全国数据
        national_df = problem_module.calculate_results(sol['genes'])

        # 保存省级和全国数据到Excel
        file_name = f"pareto_solution_{i}"
        national_path = os.path.join(PARETO_DIR, f"{file_name}_national.xlsx")
        national_df.to_excel(national_path, index=False)

        # 记录用于绘图的数据
        total_gdp = -sol['objectives'][0]  # 还原为正值（原目标是最大化GDP，存储为负值）
        total_energy = sol['objectives'][1]
        total_co2 = sol['objectives'][2]
        pareto_data.append([total_gdp, total_energy, total_co2, i])

        print(f"已保存帕累托解 {i} 到 {national_path}")

    return pareto_solutions, pareto_data


def plot_3d_pareto_front(pareto_data, problem_name):
    """绘制3D帕累托前沿并标记最优解"""
    if not pareto_data:
        print("没有帕累托解可用于绘图")
        return -1

    # 转换为numpy数组
    data = np.array(pareto_data)
    gdp_vals = data[:, 0]
    energy_vals = data[:, 1]
    co2_vals = data[:, 2]
    indices = data[:, 3].astype(int)

    # 找到"全国GDP较高且全国CO₂较低"的均衡点（最优解）
    # 标准化数据以平衡不同量级
    norm_gdp = (gdp_vals - gdp_vals.min()) / (gdp_vals.max() - gdp_vals.min())
    norm_co2 = (co2_vals - co2_vals.min()) / (co2_vals.max() - co2_vals.min())

    # 计算评分：GDP高且CO2低的解得分高
    scores = norm_gdp - norm_co2
    optimal_idx = np.argmax(scores)
    optimal_sol_index = indices[optimal_idx]

    # 创建3D图
    fig = plt.figure(figsize=(12, 10))
    ax = fig.add_subplot(111, projection='3d')

    # 绘制所有帕累托解
    scatter = ax.scatter(gdp_vals, energy_vals, co2_vals, c=scores, cmap='viridis',
                         alpha=0.7, s=100, label='帕累托前沿解')

    # 标记最优解
    ax.scatter(gdp_vals[optimal_idx], energy_vals[optimal_idx], co2_vals[optimal_idx],
               c='red', s=200, marker='*', label=f'最优解 (编号: {optimal_sol_index})')

    # 设置坐标轴标签
    ax.set_xlabel('全国累计GDP', fontsize=12)
    ax.set_ylabel('全国累计能源消耗', fontsize=12)
    ax.set_zlabel('全国累计CO₂排放', fontsize=12)

    # 添加颜色条和图例
    cbar = plt.colorbar(scatter, ax=ax)
    cbar.set_label('解决方案评分 (越高越优)', fontsize=10)
    ax.legend(fontsize=10)

    # 设置标题
    plt.title(f'{problem_name} 3D帕累托前沿', fontsize=15)

    # 保存图像
    plot_path = os.path.join(RESULT_DIR, f"{problem_name}_3d_pareto_front.png")
    plt.tight_layout()
    plt.savefig(plot_path, dpi=300)
    plt.show()

    print(f"3D帕累托前沿图已保存至: {plot_path}")
    print(f"最优解编号为: {optimal_sol_index}，对应文件: pareto_solution_{optimal_sol_index}_national.xlsx")

    return optimal_sol_index


def generate_excel_output(problem_name, all_results, national_years=range(2023, 2031)):
    """生成Excel输出文件"""
    from openpyxl.styles import Font, Alignment

    # 创建一个Excel工作簿
    wb = Workbook()
    ws_national = wb.active
    ws_national.title = "全国汇总"

    # 准备全国汇总数据
    national_data = []
    for year in national_years:
        year_data = []
        for i, result in enumerate(all_results):
            # 假设每个结果对应一个省份
            province_name = f"省份{i + 1}"
            # 提取该省份在这一年的数据
            if 'objectives' in result[0] and len(result[0]['objectives']) > 0:
                objective_value = result[0]['objectives'][0]
            else:
                objective_value = None

            year_data.append({
                '省份': province_name,
                '年份': year,
                '目标值': objective_value,
                '评估时间': result[1],
                '代数': result[2]
            })
        national_data.extend(year_data)

    # 将全国数据转换为DataFrame并写入工作表
    national_df = pd.DataFrame(national_data)
    # 将DataFrame写入工作表
    for r_idx, row in enumerate(national_df.values.tolist(), 2):
        for c_idx, value in enumerate(row, 1):
            ws_national.cell(row=r_idx, column=c_idx, value=value)

    # 添加表头
    for c_idx, col_name in enumerate(national_df.columns, 1):
        ws_national.cell(row=1, column=c_idx, value=col_name)

    # 为全国工作表添加样式
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    for cell in ws_national[1]:
        cell.font = header_font
        cell.alignment = header_alignment

    # 自动调整列宽
    for column in ws_national.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws_national.column_dimensions[column_letter].width = adjusted_width

    # 处理各省份每年的数据
    for year in national_years:
        # 为每年创建一个工作表
        ws_year = wb.create_sheet(title=str(year))

        # 准备该年的数据
        year_data = []
        for i, result in enumerate(all_results):
            province_name = f"省份{i + 1}"
            if 'objectives' in result[0] and len(result[0]['objectives']) > 0:
                objective_value = result[0]['objectives'][0]
            else:
                objective_value = None

            year_data.append({
                '省份': province_name,
                '年份': year,
                '目标值': objective_value,
                '评估时间': result[1],
                '代数': result[2]
            })

        # 将省份数据转换为DataFrame并写入工作表
        year_df = pd.DataFrame(year_data)
        for r_idx, row in enumerate(year_df.values.tolist(), 2):
            for c_idx, value in enumerate(row, 1):
                ws_year.cell(row=r_idx, column=c_idx, value=value)

        # 添加表头
        for c_idx, col_name in enumerate(year_df.columns, 1):
            ws_year.cell(row=1, column=c_idx, value=col_name)

        # 为年份工作表添加样式
        for cell in ws_year[1]:
            cell.font = header_font
            cell.alignment = header_alignment

        # 自动调整列宽
        for column in ws_year.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws_year.column_dimensions[column_letter].width = adjusted_width

    # 保存Excel文件
    excel_path = os.path.join(RESULT_DIR, f"{problem_name}_结果汇总.xlsx")
    wb.save(excel_path)
    print(f"✅ Excel结果文件已保存至: {excel_path}")

    return excel_path


if __name__ == '__main__':
    import c01

    problemModule = [c01]
    print("================================================================================")
    for m in problemModule:
        print("正在处理", m.__name__, "问题")
        problem_initialize = m.problem_initialize()
        t = 1  # 运行次数
        res, res1, res2 = [], [], []
        all_run_results = []  # 存储所有运行结果用于Excel输出

        result_path = os.path.join(RESULT_DIR, m.__name__ + ".txt")
        with open(result_path, 'w') as initFile:
            initFile.write("This is dynamic version of NSGA_II:\n")

        while t > 0:
            # 运行算法，获取结果和完整种群
            best_ind, eval_time, generations, population = run(
                problem_initialize, 100, 1000, m.evaluate)
            res.append(best_ind)
            res1.append(eval_time)
            res2.append(generations)
            all_run_results.append((best_ind, eval_time, generations))  # 保存完整结果

            with open(result_path, 'a') as f:
                f.write(f"run is {t}\n最佳目标值: {best_ind['objectives']}\n")
            t -= 1

        # 保存帕累托解
        pareto_solutions, pareto_data = save_pareto_solutions(population, m)

        # 绘制3D帕累托前沿并获取最优解
        optimal_index = plot_3d_pareto_front(pareto_data, m.__name__)

        # 记录最优解信息到结果文件
        with open(result_path, 'a') as f:
            tmp_avr = [ind['objectives'][0] for ind in res]
            ave = get_average(tmp_avr)
            var = get_variance(tmp_avr, ave)
            f.write(f"Worst:{max(tmp_avr)}\nBest:{min(tmp_avr)}\nMean:{ave}\nVaria:{var}\ngenerations:{res2}\n")
            f.write(f"最优解编号: {optimal_index}\n")
            f.write(f"最优解对应文件: pareto_solution_{optimal_index}_national.xlsx\n")

        # 生成Excel输出
        generate_excel_output(m.__name__, all_run_results)

        print(f"\n优化完成！最优解编号为: {optimal_index}")

        print("================================================================================")
